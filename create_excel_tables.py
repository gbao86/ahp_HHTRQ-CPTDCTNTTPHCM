import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_excel_with_tables():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'So sánh cặp tiêu chí'
    
    # Danh sách tiêu chí
    criteria = [
        'Chi phí',
        'Thời gian',
        'Tiện lợi',
        'An toàn',
        'Môi trường'
    ]
    
    # Style cho header
    header_font = Font(name='Arial', size=12, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Tiêu đề
    ws['A1'] = 'So sánh cặp tiêu chí'
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].alignment = center_alignment
    
    # Header
    ws['A3'] = 'Tiêu chí'
    for i, crit in enumerate(criteria, start=1):
        ws.cell(row=3, column=i+1, value=crit)
    
    # Style cho header
    for col in range(1, 7):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment
    
    # Dữ liệu mẫu cho ma trận so sánh
    sample_data = [
        [1, 3, 5, 2, 4],  # Chi phí
        [1/3, 1, 3, 1/2, 2],  # Thời gian
        [1/5, 1/3, 1, 1/4, 1/2],  # Tiện lợi
        [1/2, 2, 4, 1, 3],  # An toàn
        [1/4, 1/2, 2, 1/3, 1]  # Môi trường
    ]
    
    # Tạo ma trận so sánh với dữ liệu mẫu
    for i, crit in enumerate(criteria):
        row = i + 4
        # Tên tiêu chí ở cột đầu
        ws.cell(row=row, column=1, value=crit)
        
        # Các giá trị trong ma trận
        for j in range(len(criteria)):
            cell = ws.cell(row=row, column=j+2)
            if i == j:
                cell.value = 1  # Đường chéo chính
                cell.font = Font(bold=True)
                cell.fill = yellow_fill  # Đánh dấu đường chéo chính
            else:
                cell.value = sample_data[i][j]
            cell.border = border
            cell.alignment = center_alignment
    
    # Điều chỉnh độ rộng cột
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Thêm hướng dẫn chi tiết
    ws['A10'] = 'Hướng dẫn sử dụng:'
    ws['A11'] = '1. Ma trận phải có kích thước 5x5 (5 tiêu chí)'
    ws['A12'] = '2. Đường chéo chính (ô vàng) luôn là 1'
    ws['A13'] = '3. Nhập giá trị từ 1-9 vào các ô trống'
    ws['A14'] = '4. Giá trị nghịch đảo sẽ tự động được tính'
    ws['A15'] = '5. Dữ liệu mẫu đã được điền sẵn, bạn có thể thay đổi theo ý mình'
    ws['A16'] = '6. KHÔNG thay đổi tên các tiêu chí'
    ws['A17'] = '7. KHÔNG thay đổi cấu trúc bảng'
    
    # Style cho phần hướng dẫn
    for row in range(10, 18):
        cell = ws[f'A{row}']
        cell.font = Font(name='Arial', size=11)
        if row == 10:
            cell.font = Font(name='Arial', size=11, bold=True)
    
    # Lưu file
    wb.save('AHP_Nhap_Trong_So_Mau.xlsx')

if __name__ == "__main__":
    create_excel_with_tables() 