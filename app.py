from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
import numpy as np
import pandas as pd
import json
import MySQLdb.cursors
import matplotlib
matplotlib.use('Agg')  # Thêm dòng này vào đầu file, trước khi import plt
import matplotlib.pyplot as plt
import io
import base64
import os
import tempfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image as PILImage
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Đăng ký font Unicode hỗ trợ tiếng Việt
# Sử dụng font Arial Unicode MS có sẵn trong Windows
font_path = os.path.join(os.environ['WINDIR'], 'Fonts', 'ARIALUNI.TTF')
if os.path.exists(font_path):
    pdfmetrics.registerFont(TTFont('ArialUnicode', font_path))
    FONT_NAME = 'ArialUnicode'
else:
    # Nếu không tìm thấy Arial Unicode MS, sử dụng font mặc định
    FONT_NAME = 'Helvetica'
    print("Không tìm thấy font Arial Unicode MS, sử dụng font mặc định")

# Đăng ký font Times New Roman hỗ trợ tiếng Việt (tự động kiểm tra nhiều đường dẫn)
font_paths = [
    r'C:\Windows\Fonts\times.ttf',
    r'C:\Windows\Fonts\timesbd.ttf',
    r'C:\Windows\Fonts\Times New Roman.ttf',
    r'C:\Windows\Fonts\Times New Roman Bold.ttf'
]
font_found = False
for path in font_paths:
    if os.path.exists(path):
        pdfmetrics.registerFont(TTFont('TimesNewRoman', path))
        FONT_NAME = 'TimesNewRoman'
        font_found = True
        print(f"Đã tìm thấy font tại: {path}")
        break
if not font_found:
    FONT_NAME = 'Helvetica'
    print("Không tìm thấy font Times New Roman, sử dụng font mặc định")

app = Flask(__name__)
# Đảm bảo đây là một secret key mạnh và cố định!
app.secret_key = 'b1b9e2c3d4f5a6b7c8d9e0f1a2b3c4d5e6f7a8b9c0d1e2f3a4b5c6d7e8f9a0b1' # Ví dụ, hãy thay bằng chuỗi của bạn

# Cấu hình MySQL
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'ahp_db1'

try:
    mysql = MySQLdb.connect(
        host=app.config['MYSQL_HOST'],
        user=app.config['MYSQL_USER'],
        passwd=app.config['MYSQL_PASSWORD'],
        db=app.config['MYSQL_DB'],
        cursorclass=MySQLdb.cursors.DictCursor
    )
except MySQLdb.Error as e:
    print(f"Lỗi kết nối MySQL: {e}")
    mysql = None

# Hàm tính toán AHP (không thay đổi)
def calculate_ahp_weights(matrix_data):
    matrix = np.array(matrix_data, dtype=float)
    n = len(matrix)
    col_sums = np.sum(matrix, axis=0)
    normalized_matrix = matrix / col_sums
    weights = np.mean(normalized_matrix, axis=1)
    weighted_sum_vector = np.dot(matrix, weights)
    consistency_vector = weighted_sum_vector / weights
    lambda_max = np.mean(consistency_vector)
    CI = (lambda_max - n) / (n - 1) if n > 1 else 0
    RI_values = {
        1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32,
        8: 1.41, 9: 1.45, 10: 1.49, 11: 1.51, 12: 1.54, 13: 1.56, 14: 1.57, 15: 1.59
    }
    RI = RI_values.get(n, 1.59)
    CR = CI / RI if RI != 0 else float('inf')
    return weights.tolist(), lambda_max, CI, CR

def is_consistent(cr_value):
    return cr_value <= 0.10

def process_uploaded_file(file_path):
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path, header=None)
    elif file_path.endswith(('.xls', '.xlsx')):
        df = pd.read_excel(file_path, header=None)
    else:
        raise ValueError("Unsupported file format. Please upload CSV or Excel.")

    # Nếu có tiêu đề ở dòng đầu và cột đầu, loại bỏ chúng
    # Kiểm tra nếu phần tử đầu tiên không phải là số, coi là tiêu đề
    if not pd.api.types.is_numeric_dtype(df.iloc[0, 1:]) or not pd.api.types.is_numeric_dtype(df.iloc[1:, 0]):
        df = df.iloc[1:, 1:]
    # Nếu vẫn còn cột/dòng không phải số, tiếp tục loại bỏ
    while df.shape[0] > 0 and not pd.api.types.is_numeric_dtype(df.iloc[0]):
        df = df.iloc[1:]
    while df.shape[1] > 0 and not pd.api.types.is_numeric_dtype(df.iloc[:,0]):
        df = df.iloc[:,1:]

    # Chuyển về float và trả về list
    df = df.apply(pd.to_numeric, errors='coerce')
    df = df.dropna(axis=0, how='any')
    df = df.dropna(axis=1, how='any')
    return df.values.tolist()

@app.route('/')
def index():
    session.clear()
    return render_template('index.html', current_year=2025)

@app.route('/criteria_comparison', methods=['GET', 'POST'])
def criteria_comparison():
    criteria = ['Chi phí', 'Thời gian', 'Tiện lợi', 'An toàn', 'Môi trường']
    session['criteria'] = criteria

    if request.method == 'POST':
        if 'file_upload' in request.files:
            file = request.files['file_upload']
            if file and file.filename != '':
                try:
                    filepath = f"temp_{file.filename}"
                    file.save(filepath)
                    matrix_data = process_uploaded_file(filepath)
                    print("DEBUG matrix_data:", matrix_data)  # Thêm dòng này để debug
                    if len(matrix_data) != len(criteria) or len(matrix_data[0]) != len(criteria):
                        return render_template('criteria_comparison.html', criteria=criteria, error="Kích thước ma trận trong file không khớp với số lượng tiêu chí.", current_year=2025)
                except Exception as e:
                    return render_template('criteria_comparison.html', criteria=criteria, error=f"Lỗi khi đọc file: {e}", current_year=2025)
                finally:
                    if os.path.exists(filepath):
                        os.remove(filepath)
            else:
                return render_template('criteria_comparison.html', criteria=criteria, error="Vui lòng chọn file hợp lệ.", current_year=2025)
        else: # Nhập thủ công
            matrix_data = []
            for i in range(len(criteria)):
                row = []
                for j in range(len(criteria)):
                    try:
                        val = float(request.form[f'criteria_{i}_{j}'])
                        row.append(val)
                    except ValueError:
                        return render_template('criteria_comparison.html', criteria=criteria, error="Giá trị nhập không hợp lệ.", current_year=2025)
                matrix_data.append(row)

        try:
            criteria_weights, lambda_max, CI, CR = calculate_ahp_weights(matrix_data)

            session['criteria_matrix'] = matrix_data
            session['criteria_weights'] = criteria_weights
            session['criteria_lambda_max'] = lambda_max
            session['criteria_CI'] = CI
            session['criteria_CR'] = CR

            if is_consistent(CR):
                return render_template('criteria_comparison.html',
                                       criteria=criteria,
                                       matrix_data=matrix_data,
                                       weights=criteria_weights,
                                       lambda_max=lambda_max,
                                       CI=CI,
                                       CR=CR,
                                       is_consistent=True,
                                       current_year=2025)
            else:
                return render_template('criteria_comparison.html',
                                       criteria=criteria,
                                       matrix_data=matrix_data,
                                       weights=criteria_weights,
                                       lambda_max=lambda_max,
                                       CI=CI,
                                       CR=CR,
                                       is_consistent=False,
                                       error="Tỷ số nhất quán CR lớn hơn 10%. Vui lòng nhập lại.",
                                       current_year=2025)
        except Exception as e:
            return render_template('criteria_comparison.html', criteria=criteria, error=f"Có lỗi xảy ra trong quá trình tính toán: {e}", current_year=2025)

    return render_template('criteria_comparison.html', criteria=criteria, current_year=2025)

# Trang so sánh phương án theo từng tiêu chí
@app.route('/alternatives_comparison', methods=['GET', 'POST'])
def alternatives_comparison():
    if 'criteria_weights' not in session:
        return redirect(url_for('criteria_comparison'))

    criteria = session.get('criteria', [])

    if request.method == 'POST':
        # --- Lấy dữ liệu số lượng và tên phương án từ hidden inputs (do JS cập nhật) ---
        try:
            num_alts = int(request.form['num_alternatives'])
            if not (3 <= num_alts <= 5):
                raise ValueError("Số lượng phương án phải từ 3 đến 5.")
            
            new_alternatives = []
            for i in range(num_alts):
                # Lấy tên phương án từ request.form (từ các hidden input)
                alt_name = request.form.get(f'alt_name_{i}', f'Phương án {i+1}')
                new_alternatives.append(alt_name if alt_name else f'Phương án {i+1}')
            
            session['num_alternatives'] = num_alts
            session['alternatives'] = new_alternatives
            
        except ValueError as e:
            # Nếu có lỗi khi lấy num_alternatives hoặc tên, quay lại trang với thông báo lỗi
            return render_template('alternatives_comparison.html',
                                   criteria=criteria,
                                   num_alternatives=session.get('num_alternatives', 3), # Dùng giá trị hiện tại trong session
                                   alternatives=session.get('alternatives', [f'Phương án {i+1}' for i in range(session.get('num_alternatives', 3))]),
                                   error_alt=str(e),
                                   current_year=2025)
        
        # --- Tiếp tục xử lý ma trận so sánh cho tiêu chí hiện tại ---
        alternatives = session.get('alternatives', []) # Lấy danh sách phương án đã cập nhật từ session
        num_alternatives = len(alternatives) # Đảm bảo số lượng phương án khớp với danh sách lấy được

        alternative_weights_by_criteria = session.get('alternative_weights_by_criteria', {})
        processed_criteria_count = len(alternative_weights_by_criteria)
        current_criteria_index = processed_criteria_count

        if current_criteria_index < len(criteria):
            current_criteria_name = criteria[current_criteria_index]

            matrix_data = []
            file_uploaded_successfully = False # Biến cờ để kiểm tra xem có file nào được tải lên thành công không
            
            if 'file_upload' in request.files:
                file = request.files['file_upload']
                if file and file.filename != '': # Chỉ xử lý nếu CÓ file được chọn và CÓ tên file
                    file_uploaded_successfully = True
                    try:
                        filepath = f"temp_{file.filename}"
                        file.save(filepath)
                        matrix_data = process_uploaded_file(filepath)
                        if len(matrix_data) != num_alternatives or len(matrix_data[0]) != num_alternatives:
                            return render_template('alternatives_comparison.html',
                                                   criteria=criteria,
                                                   num_alternatives=num_alternatives,
                                                   alternatives=alternatives,
                                                   current_criteria_index=current_criteria_index,
                                                   current_criteria_name=current_criteria_name,
                                                   error=f"Kích thước ma trận trong file không khớp với số lượng phương án ({num_alternatives}).",
                                                   current_year=2025)
                    except Exception as e:
                        return render_template('alternatives_comparison.html',
                                               criteria=criteria,
                                               num_alternatives=num_alternatives,
                                               alternatives=alternatives,
                                               current_criteria_index=current_criteria_index,
                                               current_criteria_name=current_criteria_name,
                                               error=f"Lỗi khi đọc file: {e}",
                                               current_year=2025)
                    finally:
                        if os.path.exists(filepath):
                            os.remove(filepath)
            
            # Nếu không có file nào được tải lên thành công, thì xử lý nhập liệu thủ công
            if not file_uploaded_successfully:
                for i in range(num_alternatives):
                    row = []
                    for j in range(num_alternatives):
                        try:
                            # Sử dụng request.form.get() với giá trị mặc định để tránh lỗi KeyError nếu ô trống
                            val = float(request.form.get(f'alt_{i}_{j}', '0')) # Mặc định là '0' nếu không có giá trị
                            row.append(val)
                        except ValueError:
                            return render_template('alternatives_comparison.html',
                                                   criteria=criteria,
                                                   num_alternatives=num_alternatives,
                                                   alternatives=alternatives,
                                                   current_criteria_index=current_criteria_index,
                                                   current_criteria_name=current_criteria_name,
                                                   error="Giá trị nhập không hợp lệ.",
                                                   current_year=2025)
                    matrix_data.append(row)

            try:
                alt_weights, alt_lambda_max, alt_CI, alt_CR = calculate_ahp_weights(matrix_data)

                if is_consistent(alt_CR):
                    alternative_weights_by_criteria[current_criteria_name] = {
                        'matrix': matrix_data,
                        'weights': alt_weights,
                        'lambda_max': alt_lambda_max,
                        'CI': alt_CI,
                        'CR': alt_CR
                    }
                    session['alternative_weights_by_criteria'] = alternative_weights_by_criteria
                    # Chuyển sang tiêu chí tiếp theo hoặc trang kết quả
                    if current_criteria_index + 1 < len(criteria):
                        return redirect(url_for('alternatives_comparison'))
                    else:
                        return redirect(url_for('results'))
                else:
                    # Nếu không nhất quán, render lại trang hiện tại với lỗi và dữ liệu tạm thời
                    return render_template('alternatives_comparison.html',
                                           criteria=criteria,
                                           num_alternatives=num_alternatives,
                                           alternatives=alternatives,
                                           current_criteria_index=current_criteria_index,
                                           current_criteria_name=current_criteria_name,
                                           matrix_data=matrix_data,
                                           weights=alt_weights,
                                           lambda_max=alt_lambda_max,
                                           CI=alt_CI,
                                           CR=alt_CR,
                                           is_consistent=False,
                                           error="Tỷ số nhất quán CR lớn hơn 10%. Vui lòng nhập lại.",
                                           current_year=2025)
            except Exception as e:
                return render_template('alternatives_comparison.html',
                                       criteria=criteria,
                                       num_alternatives=num_alternatives,
                                       alternatives=alternatives,
                                       current_criteria_index=current_criteria_index,
                                       current_criteria_name=current_criteria_name,
                                       error=f"Có lỗi xảy ra trong quá trình tính toán: {e}",
                                       current_year=2025)
        else: # current_criteria_index >= len(criteria)
            # Trường hợp đã xử lý hết tất cả các tiêu chí
            return redirect(url_for('results'))


    # Lần tải trang đầu tiên (GET request)
    # Khởi tạo num_alternatives và alternatives từ session hoặc giá trị mặc định
    num_alternatives = session.get('num_alternatives')
    alternatives = session.get('alternatives')

    if num_alternatives is None:
        num_alternatives = 3 # Giá trị mặc định ban đầu
        alternatives = [f'Phương án {i+1}' for i in range(num_alternatives)]
        session['num_alternatives'] = num_alternatives
        session['alternatives'] = alternatives
        # Không cần redirect ở đây vì JS sẽ handle initial display

    current_criteria_index = len(session.get('alternative_weights_by_criteria', {}))
    current_criteria_name = criteria[current_criteria_index] if current_criteria_index < len(criteria) else None

    return render_template('alternatives_comparison.html',
                           criteria=criteria,
                           num_alternatives=num_alternatives,
                           alternatives=alternatives,
                           current_criteria_index=current_criteria_index,
                           current_criteria_name=current_criteria_name,
                           current_year=2025)

# Trang kết quả (không thay đổi)
@app.route('/results')
def results():
    criteria_weights = session.get('criteria_weights')
    alternative_weights_by_criteria = session.get('alternative_weights_by_criteria')
    criteria_names = session.get('criteria')
    alternative_names = session.get('alternatives')

    if not criteria_weights or not alternative_weights_by_criteria or not criteria_names or not alternative_names:
        return redirect(url_for('criteria_comparison'))

    final_alternative_scores = np.zeros(len(alternative_names))
    for i, alt_name in enumerate(alternative_names):
        for j, crit_name in enumerate(criteria_names):
            weight_alt_by_crit = alternative_weights_by_criteria[crit_name]['weights'][i]
            weight_criteria = criteria_weights[j]
            final_alternative_scores[i] += weight_alt_by_crit * weight_criteria

    best_alternative_index = np.argmax(final_alternative_scores)
    best_alternative_name = alternative_names[best_alternative_index]

    plt.rcParams.update({'font.size': 10})
    fig_pie, ax_pie = plt.subplots(figsize=(7, 7))
    ax_pie.pie(criteria_weights, labels=criteria_names, autopct='%1.1f%%', startangle=90, pctdistance=0.85)
    ax_pie.axis('equal')
    ax_pie.set_title('Trọng số các tiêu chí')
    pie_chart_buffer = io.BytesIO()
    plt.savefig(pie_chart_buffer, format='png', bbox_inches='tight')
    pie_chart_buffer.seek(0)
    pie_chart_base64 = base64.b64encode(pie_chart_buffer.getvalue()).decode('utf-8')
    plt.close(fig_pie)

    fig_bar, ax_bar = plt.subplots(figsize=(10, 6))
    bars = ax_bar.bar(alternative_names, final_alternative_scores, color='skyblue')
    ax_bar.set_ylabel('Điểm')
    ax_bar.set_title('Điểm Số Các Phương Án')
    ax_bar.tick_params(axis='x', rotation=45)
    for bar in bars:
        yval = bar.get_height()
        ax_bar.text(bar.get_x() + bar.get_width()/2, yval + 0.005, round(yval, 3), ha='center', va='bottom')
    plt.tight_layout()
    bar_chart_buffer = io.BytesIO()
    plt.savefig(bar_chart_buffer, format='png', bbox_inches='tight')
    bar_chart_buffer.seek(0)
    bar_chart_base64 = base64.b64encode(bar_chart_buffer.getvalue()).decode('utf-8')
    plt.close(fig_bar)

    if mysql:
        try:
            cursor = mysql.cursor()
            cursor.execute(
                "INSERT INTO ahp_history (goal_name, criteria_weights, alternatives_results, chosen_alternative, visual_data, alt_weights_detail) VALUES (%s, %s, %s, %s, %s, %s)",
                (
                    "Chọn Phương Tiện Di Chuyển Trong Nội Thành TPHCM",
                    json.dumps({c: w for c, w in zip(criteria_names, criteria_weights)}),
                    json.dumps({alt: score for alt, score in zip(alternative_names, final_alternative_scores.tolist())}),
                    best_alternative_name,
                    json.dumps({
                        'pie_chart': pie_chart_base64,
                        'bar_chart': bar_chart_base64
                    }),
                    json.dumps(alternative_weights_by_criteria)
                )
            )
            mysql.commit()
        except Exception as e:
            print(f"Lỗi khi lưu vào DB: {e}")
    else:
        print("Không thể lưu kết quả vào DB vì kết nối MySQL không thành công.")

    detailed_alternative_results = []
    for i, alt_name in enumerate(alternative_names):
        detail_row = {'alternative': alt_name, 'final_score': final_alternative_scores[i]}
        for j, crit_name in enumerate(criteria_names):
            detail_row[crit_name] = alternative_weights_by_criteria[crit_name]['weights'][i]
        detailed_alternative_results.append(detail_row)

    return render_template('results.html',
                           criteria_names=criteria_names,
                           criteria_weights={c: w for c, w in zip(criteria_names, criteria_weights)},
                           alternative_names=alternative_names,
                           final_alternative_scores={alt: score for alt, score in zip(alternative_names, final_alternative_scores.tolist())},
                           best_alternative_name=best_alternative_name,
                           pie_chart_base64=pie_chart_base64,
                           bar_chart_base64=bar_chart_base64,
                           detailed_alternative_results=detailed_alternative_results,
                           current_year=2025)

@app.route('/history')
def history():
    records = []
    if mysql:
        try:
            cursor = mysql.cursor()
            cursor.execute("SELECT * FROM ahp_history ORDER BY timestamp DESC")
            records = cursor.fetchall()

            for record in records:
                record['criteria_weights'] = json.loads(record['criteria_weights'])
                record['alternatives_results'] = json.loads(record['alternatives_results'])
                record['visual_data'] = json.loads(record['visual_data'])
        except Exception as e:
            records = []
            print(f"Lỗi khi lấy lịch sử từ DB: {e}")
    else:
        print("Không thể lấy lịch sử từ DB vì kết nối MySQL không thành công.")

    return render_template('history.html', records=records, current_year=2025)

@app.route('/export_excel')
def export_excel():
    record_id = request.args.get('record_id')
    if record_id and mysql:
        try:
            cursor = mysql.cursor()
            cursor.execute("SELECT * FROM ahp_history WHERE id = %s", (record_id,))
            record = cursor.fetchone()
            if record:
                criteria_weights = json.loads(record['criteria_weights'])
                alternatives_results = json.loads(record['alternatives_results'])
                criteria_names = list(criteria_weights.keys())
                alternative_names = list(alternatives_results.keys())
                final_alternative_scores = [alternatives_results[alt] for alt in alternative_names]
                best_alternative_name = record['chosen_alternative']
                # Lấy dữ liệu chi tiết nếu có
                alt_weights_detail = json.loads(record['alt_weights_detail']) if record.get('alt_weights_detail') else None
            else:
                return redirect(url_for('criteria_comparison'))
        except Exception as e:
            print(f"Lỗi khi lấy dữ liệu từ DB: {e}")
            return redirect(url_for('criteria_comparison'))
    else:
        criteria_weights = session.get('criteria_weights')
        alternative_weights_by_criteria = session.get('alternative_weights_by_criteria')
        criteria_names = session.get('criteria')
        alternative_names = session.get('alternatives')
        best_alternative_name = None
        if not criteria_weights or not alternative_weights_by_criteria or not criteria_names or not alternative_names:
            return redirect(url_for('criteria_comparison'))

        final_alternative_scores = np.zeros(len(alternative_names))
        for i, alt_name in enumerate(alternative_names):
            for j, crit_name in enumerate(criteria_names):
                weight_alt_by_crit = alternative_weights_by_criteria[crit_name]['weights'][i]
                weight_criteria = criteria_weights[j]
                final_alternative_scores[i] += weight_alt_by_crit * weight_criteria
        best_alternative_index = np.argmax(final_alternative_scores)
        best_alternative_name = alternative_names[best_alternative_index]
        alt_weights_detail = alternative_weights_by_criteria

    # Prepare DataFrames
    if isinstance(criteria_weights, dict):
        weights_list = [criteria_weights[crit] for crit in criteria_names]
    else:
        weights_list = criteria_weights

    df_criteria = pd.DataFrame({
        'Tiêu chí': criteria_names,
        'Trọng số': weights_list
    })
    df_alternatives = pd.DataFrame({
        'Phương án': alternative_names,
        'Điểm cuối cùng': final_alternative_scores
    })
    # Detailed table
    detail_data = []
    for i, alt_name in enumerate(alternative_names):
        row = {'Phương án': alt_name, 'Điểm cuối cùng': final_alternative_scores[i]}
        for crit_name in criteria_names:
            if alt_weights_detail and crit_name in alt_weights_detail:
                try:
                    # Tìm index của phương án trong danh sách alternative_names
                    alt_idx = alternative_names.index(alt_name)
                    row[f'Trọng số theo {crit_name}'] = alt_weights_detail[crit_name]['weights'][alt_idx]
                except Exception:
                    row[f'Trọng số theo {crit_name}'] = ''
            else:
                row[f'Trọng số theo {crit_name}'] = ''
        detail_data.append(row)
    df_detail = pd.DataFrame(detail_data)

    # Generate charts again (as in /results)
    plt.rcParams.update({'font.size': 10})
    fig_pie, ax_pie = plt.subplots(figsize=(7, 7))
    ax_pie.pie(weights_list, labels=criteria_names, autopct='%1.1f%%', startangle=90, pctdistance=0.85)
    ax_pie.axis('equal')
    ax_pie.set_title('Trọng số các tiêu chí')
    pie_chart_buffer = io.BytesIO()
    plt.savefig(pie_chart_buffer, format='png', bbox_inches='tight')
    pie_chart_buffer.seek(0)
    plt.close(fig_pie)

    fig_bar, ax_bar = plt.subplots(figsize=(10, 6))
    bars = ax_bar.bar(alternative_names, final_alternative_scores, color='skyblue')
    ax_bar.set_ylabel('Điểm')
    ax_bar.set_title('Điểm Số Các Phương Án')
    ax_bar.tick_params(axis='x', rotation=45)
    for bar in bars:
        yval = bar.get_height()
        ax_bar.text(bar.get_x() + bar.get_width()/2, yval + 0.005, round(yval, 3), ha='center', va='bottom')
    plt.tight_layout()
    bar_chart_buffer = io.BytesIO()
    plt.savefig(bar_chart_buffer, format='png', bbox_inches='tight')
    bar_chart_buffer.seek(0)
    plt.close(fig_bar)

    # Create Excel file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Trọng số tiêu chí'
        for r in dataframe_to_rows(df_criteria, index=False, header=True):
            ws1.append(r)
        ws2 = wb.create_sheet('Điểm phương án')
        for r in dataframe_to_rows(df_alternatives, index=False, header=True):
            ws2.append(r)
        ws3 = wb.create_sheet('Bảng chi tiết')
        for r in dataframe_to_rows(df_detail, index=False, header=True):
            ws3.append(r)
        # Add charts as images
        # Pie chart
        pie_img = PILImage.open(pie_chart_buffer)
        pie_img_path = tempfile.mktemp(suffix='.png')
        pie_img.save(pie_img_path)
        img1 = XLImage(pie_img_path)
        ws1.add_image(img1, 'E2')
        # Bar chart
        bar_img = PILImage.open(bar_chart_buffer)
        bar_img_path = tempfile.mktemp(suffix='.png')
        bar_img.save(bar_img_path)
        img2 = XLImage(bar_img_path)
        ws2.add_image(img2, 'E2')
        # Summary sheet
        ws_sum = wb.create_sheet('Tóm tắt')
        ws_sum['A1'] = 'Phân tích AHP - Kết quả cuối cùng'
        ws_sum['A2'] = f'Phương án được chọn: {best_alternative_name}'
        ws_sum['A3'] = 'Tiêu chí:'
        for idx, crit in enumerate(criteria_names):
            ws_sum[f'A{4+idx}'] = f'- {crit}: {weights_list[idx]:.4f}'
        ws_sum[f'A{4+len(criteria_names)}'] = 'Điểm các phương án:'
        for idx, alt in enumerate(alternative_names):
            ws_sum[f'A{5+len(criteria_names)+idx}'] = f'- {alt}: {final_alternative_scores[idx]:.4f}'
        wb.save(tmp.name)
        # Clean up temp images
        os.remove(pie_img_path)
        os.remove(bar_img_path)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name='AHP_Ket_Qua.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export_pdf')
def export_pdf():
    record_id = request.args.get('record_id')
    if record_id and mysql:
        try:
            cursor = mysql.cursor()
            cursor.execute("SELECT * FROM ahp_history WHERE id = %s", (record_id,))
            record = cursor.fetchone()
            if record:
                criteria_weights = json.loads(record['criteria_weights'])
                alternatives_results = json.loads(record['alternatives_results'])
                criteria_names = list(criteria_weights.keys())
                alternative_names = list(alternatives_results.keys())
                final_alternative_scores = [alternatives_results[alt] for alt in alternative_names]
                best_alternative_name = record['chosen_alternative']
                # Lấy dữ liệu chi tiết nếu có
                alt_weights_detail = json.loads(record['alt_weights_detail']) if record.get('alt_weights_detail') else None
            else:
                return redirect(url_for('criteria_comparison'))
        except Exception as e:
            print(f"Lỗi khi lấy dữ liệu từ DB: {e}")
            return redirect(url_for('criteria_comparison'))
    else:
        criteria_weights = session.get('criteria_weights')
        alternative_weights_by_criteria = session.get('alternative_weights_by_criteria')
        criteria_names = session.get('criteria')
        alternative_names = session.get('alternatives')
        best_alternative_name = None
        if not criteria_weights or not alternative_weights_by_criteria or not criteria_names or not alternative_names:
            return redirect(url_for('criteria_comparison'))

        final_alternative_scores = np.zeros(len(alternative_names))
        for i, alt_name in enumerate(alternative_names):
            for j, crit_name in enumerate(criteria_names):
                weight_alt_by_crit = alternative_weights_by_criteria[crit_name]['weights'][i]
                weight_criteria = criteria_weights[j]
                final_alternative_scores[i] += weight_alt_by_crit * weight_criteria
        best_alternative_index = np.argmax(final_alternative_scores)
        best_alternative_name = alternative_names[best_alternative_index]
        alt_weights_detail = alternative_weights_by_criteria

    # Prepare DataFrames
    if isinstance(criteria_weights, dict):
        weights_list = [criteria_weights[crit] for crit in criteria_names]
    else:
        weights_list = criteria_weights

    df_criteria = pd.DataFrame({
        'Tiêu chí': criteria_names,
        'Trọng số': weights_list
    })
    df_alternatives = pd.DataFrame({
        'Phương án': alternative_names,
        'Điểm cuối cùng': final_alternative_scores
    })
    # Detailed table
    detail_data = []
    for i, alt_name in enumerate(alternative_names):
        row = {'Phương án': alt_name, 'Điểm cuối cùng': final_alternative_scores[i]}
        for crit_name in criteria_names:
            if alt_weights_detail and crit_name in alt_weights_detail:
                try:
                    # Tìm index của phương án trong danh sách alternative_names
                    alt_idx = alternative_names.index(alt_name)
                    row[f'Trọng số theo {crit_name}'] = alt_weights_detail[crit_name]['weights'][alt_idx]
                except Exception:
                    row[f'Trọng số theo {crit_name}'] = ''
            else:
                row[f'Trọng số theo {crit_name}'] = ''
        detail_data.append(row)
    df_detail = pd.DataFrame(detail_data)

    # Generate charts again (as in /results)
    plt.rcParams.update({'font.size': 10})
    fig_pie, ax_pie = plt.subplots(figsize=(7, 7))
    ax_pie.pie(weights_list, labels=criteria_names, autopct='%1.1f%%', startangle=90, pctdistance=0.85)
    ax_pie.axis('equal')
    ax_pie.set_title('Trọng số các tiêu chí')
    pie_chart_buffer = io.BytesIO()
    plt.savefig(pie_chart_buffer, format='png', bbox_inches='tight')
    pie_chart_buffer.seek(0)
    plt.close(fig_pie)

    fig_bar, ax_bar = plt.subplots(figsize=(10, 6))
    bars = ax_bar.bar(alternative_names, final_alternative_scores, color='skyblue')
    ax_bar.set_ylabel('Điểm')
    ax_bar.set_title('Điểm Số Các Phương Án')
    ax_bar.tick_params(axis='x', rotation=45)
    for bar in bars:
        yval = bar.get_height()
        ax_bar.text(bar.get_x() + bar.get_width()/2, yval + 0.005, round(yval, 3), ha='center', va='bottom')
    plt.tight_layout()
    bar_chart_buffer = io.BytesIO()
    plt.savefig(bar_chart_buffer, format='png', bbox_inches='tight')
    bar_chart_buffer.seek(0)
    plt.close(fig_bar)

    # Create PDF file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        doc = SimpleDocTemplate(tmp.name, pagesize=letter)
        styles = getSampleStyleSheet()
        
        # Tạo style mới với font Unicode
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=FONT_NAME,
            fontSize=16,
            spaceAfter=30
        )
        
        heading2_style = ParagraphStyle(
            'CustomHeading2',
            parent=styles['Heading2'],
            fontName=FONT_NAME,
            fontSize=14,
            spaceAfter=12
        )
        
        heading3_style = ParagraphStyle(
            'CustomHeading3',
            parent=styles['Heading3'],
            fontName=FONT_NAME,
            fontSize=12,
            spaceAfter=12
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=FONT_NAME,
            fontSize=10
        )

        elements = []

        # Title
        elements.append(Paragraph('Phân tích AHP - Kết quả cuối cùng', title_style))
        elements.append(Paragraph(f'Phương án được chọn: {best_alternative_name}', heading2_style))
        elements.append(Spacer(1, 12))

        # Criteria weights table
        elements.append(Paragraph('Trọng số các Tiêu chí:', heading3_style))
        criteria_data = [['Tiêu chí', 'Trọng số']] + [
            [crit, f'{(criteria_weights[crit] if isinstance(criteria_weights, dict) else criteria_weights[i]):.4f}']
            for i, crit in enumerate(criteria_names)
        ]
        criteria_table = Table(criteria_data)
        criteria_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(criteria_table)
        elements.append(Spacer(1, 12))

        # Alternatives scores table
        elements.append(Paragraph('Điểm số tổng thể của các Phương án:', heading3_style))
        alt_data = [['Phương án', 'Điểm cuối cùng']] + [[alt, f'{score:.4f}'] for alt, score in zip(alternative_names, final_alternative_scores)]
        alt_table = Table(alt_data)
        alt_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(alt_table)
        elements.append(Spacer(1, 12))

        # Detailed table
        elements.append(Paragraph('Bảng Kết Quả Chi Tiết Phương Án:', heading3_style))
        # Sử dụng Paragraph cho tiêu đề cột để tự xuống dòng
        detail_headers = [
            Paragraph('Phương án', normal_style)
        ] + [
            Paragraph(f'Trọng số<br/>{crit}', normal_style) for crit in criteria_names
        ] + [
            Paragraph('Điểm<br/>cuối cùng', normal_style)
        ]
        detail_data = [detail_headers]
        for row in df_detail.to_dict('records'):
            detail_row = [row['Phương án']]
            for crit in criteria_names:
                detail_row.append(f'{row[f"Trọng số theo {crit}"]:.4f}' if row[f"Trọng số theo {crit}"] != '' else '')
            detail_row.append(f'{row["Điểm cuối cùng"]:.4f}')
            detail_data.append(detail_row)
        # Đặt colWidths hợp lý: cột đầu 60, mỗi cột trọng số 80, cột cuối 70
        col_widths = [60] + [80]*len(criteria_names) + [70]
        detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(detail_table)
        elements.append(Spacer(1, 12))

        # Add charts as images
        pie_img = PILImage.open(pie_chart_buffer)
        pie_img_path = tempfile.mktemp(suffix='.png')
        pie_img.save(pie_img_path)
        elements.append(Paragraph('Biểu đồ Trọng số Tiêu chí:', heading3_style))
        elements.append(Image(pie_img_path, width=400, height=400))
        elements.append(Spacer(1, 12))

        bar_img = PILImage.open(bar_chart_buffer)
        bar_img_path = tempfile.mktemp(suffix='.png')
        bar_img.save(bar_img_path)
        elements.append(Paragraph('Biểu đồ Điểm Các Phương án:', heading3_style))
        elements.append(Image(bar_img_path, width=500, height=300))
        elements.append(Spacer(1, 12))

        # Build PDF
        doc.build(elements)
        # Clean up temp images
        os.remove(pie_img_path)
        os.remove(bar_img_path)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name='AHP_Ket_Qua.pdf', mimetype='application/pdf')

@app.route('/delete_history/<int:record_id>', methods=['POST'])
def delete_history(record_id):
    if mysql:
        try:
            cursor = mysql.cursor()
            cursor.execute("DELETE FROM ahp_history WHERE id = %s", (record_id,))
            mysql.commit()
            flash('Đã xóa bản ghi lịch sử thành công!', 'success')
        except Exception as e:
            print(f"Lỗi khi xóa bản ghi lịch sử: {e}")
            flash('Lỗi khi xóa bản ghi lịch sử!', 'error')
    else:
        flash('Không thể kết nối MySQL!', 'error')
    return redirect(url_for('history'))

if __name__ == '__main__':
    app.run(debug=True)